# C:\Users\yyyyw\Desktop\script\팀과제\기상청41_단기예보 조회서비스_오픈API활용가이드_격자_위경도(20220103).xlsx
# import pandas as pd
from asyncio.windows_events import NULL
import openpyxl
from collections import defaultdict # 초기화를 해주지 않아도 된다.

# df = pd.read_excel(r'C:\Users\yyyyw\Desktop\script\팀과제\기상청41_단기예보 조회서비스_오픈API활용가이드_격자_위경도(20220103).xlsx')
# 
# print(df)

wb =openpyxl.load_workbook(r'C:\Users\yyyyw\Desktop\script\팀과제\기상청41_단기예보 조회서비스_오픈API활용가이드_격자_위경도(20220103).xlsx')

sheet = wb.get_sheet_by_name('최종 업데이트 파일_20220103')
data1 = sheet['A1'].value
data2 = sheet.cell(row=3, column=1).value
# row = 세로 이게 2부터 3789 까지 돌면서 검사해야함.
# column = 3 : 시(1단계), 4 : 구(2단계) 5: 동(3단계) 6 : 격자 X, 7 : 격자 Y ... 엑셀 파일 참고.
# 14 : 경도, 15 : 위도 이므로 15가 앞에 와야한다.
adress = defaultdict(dict)


for r in range(2, 3789+1):
    level_1 = sheet.cell(row=r, column=3).value
    level_2 = sheet.cell(row=r, column=4).value
    level_3 = sheet.cell(row=r, column=5).value
    X = sheet.cell(row=r, column=6).value
    Y = sheet.cell(row=r, column=7).value
    latitude = sheet.cell(row=r, column=15).value
    longitude = sheet.cell(row=r, column=14).value

    # if level_2 == '':
        # adress[level_1] = defaultdict(dict)
        # adress[level_1][level_1] = defaultdict(dict)
        # adress[level_1][level_1][level_1] =(X, Y, latitude, longitude)
        # continue
    # if level_3 == '':
        # adress[level_1][level_2] = defaultdict(dict)
        # adress[level_1][level_2][level_2] =(X, Y, latitude, longitude)
        # continue
    # adress[level_1][level_2][level_3] = (X, Y, latitude, longitude)

    if level_2 == '':
        adress['level_1'][level_1] =(X, Y, latitude, longitude)
        continue
    if level_3 == '':
        adress['level_2'][level_2] =(X, Y, latitude, longitude)
        continue
    adress['level_3'][level_3] = (X, Y, latitude, longitude)

import pickle

f = open('adress', 'wb') #pickle 사용을 위해 바이너리 쓰기 파일 오픈
pickle.dump(adress, f) #리스트 객체를 파일로 dump
f.close()

# import pprint
# pprint.pprint(adress)

f = open('adress', 'rb') #pickle 사용을 위해 바이너리 읽기 파일 오픈
dict = pickle.load(f) #파일에서 리스트 load

import pprint
pprint.pprint(dict)